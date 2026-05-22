# app.py - QTC Smart Sales Pro v5.0 (COMPLETO)

import streamlit as st
import pandas as pd
import re
import io
from datetime import datetime
from difflib import SequenceMatcher

st.set_page_config(
    page_title="QTC Smart Sales Pro",
    page_icon="💼",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ============================================
# CSS COMPLETO
# ============================================

st.markdown("""
<style>
    .stApp { background: linear-gradient(135deg, #0d47a1 0%, #1565c0 50%, #1e88e5 100%); }
    [data-testid="stSidebar"] { background: linear-gradient(180deg, #f8a35e 0%, #e87a2d 50%, #d45a1a 100%); }
    [data-testid="stSidebar"] * { color: #ffffff !important; }
    .badge-yessica { background: #4CAF50; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.7rem; display: inline-block; margin: 2px; }
    .badge-apri004 { background: #FF9800; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.7rem; display: inline-block; margin: 2px; }
    .badge-apri001 { background: #f44336; color: white; padding: 4px 12px; border-radius: 20px; font-size: 0.7rem; display: inline-block; margin: 2px; }
    .footer { text-align: center; padding: 1rem; color: rgba(255,255,255,0.7); font-size: 0.7rem; margin-top: 2rem; }
    div[style*="background:white"] * { color: #1a1a2e !important; }
    .counter-summary { background: rgba(0,0,0,0.3); border-radius: 12px; padding: 1rem; margin-bottom: 1rem; display: flex; justify-content: space-around; flex-wrap: wrap; }
    .counter-item { text-align: center; padding: 0.5rem; }
</style>
""", unsafe_allow_html=True)

# ============================================
# FUNCIONES PRINCIPALES
# ============================================

def corregir_numero(valor) -> float:
    if pd.isna(valor) or str(valor).strip() in ["", "0", "0.0", "-"]:
        return 0.0
    s = str(valor).upper().replace('S/', '').replace('$', '').replace(' ', '').strip()
    if ',' in s and '.' in s:
        s = s.replace(',', '')
    elif ',' in s:
        partes = s.split(',')
        if len(partes[-1]) <= 2:
            s = s.replace(',', '.')
        else:
            s = s.replace(',', '')
    s = re.sub(r'[^\d.]', '', s)
    try:
        return float(s)
    except:
        return 0.0

def limpiar_cabeceras(df: pd.DataFrame) -> pd.DataFrame:
    for i in range(min(20, len(df))):
        fila = [str(x).upper() for x in df.iloc[i].values]
        if any(h in item for h in ['SKU', 'COD', 'SAP', 'NUMERO', 'ARTICULO'] for item in fila):
            df.columns = [str(c).strip() for c in df.iloc[i]]
            return df.iloc[i+1:].reset_index(drop=True)
    return df

def detectar_columna_sku(df: pd.DataFrame) -> str:
    posibles = ['SKU', 'COD', 'SAP', 'NUMERO', 'ARTICULO', 'CODIGO']
    for col in df.columns:
        col_upper = str(col).upper()
        for posible in posibles:
            if posible.upper() in col_upper:
                return col
    return df.columns[0]

def detectar_columna_descripcion(df: pd.DataFrame) -> str:
    posibles = ['DESC', 'DESCRIPCION', 'NOMBRE', 'PRODUCTO', 'GOODS']
    for col in df.columns:
        col_upper = str(col).upper()
        for posible in posibles:
            if posible.upper() in col_upper:
                return col
    return None

def detectar_columnas_precio(df: pd.DataFrame) -> dict:
    precios = {}
    mapeo = {'P. IR': ['IR', 'MAYORISTA', 'MAYOR'], 
             'P. BOX': ['BOX', 'CAJA'], 
             'P. VIP': ['VIP']}
    
    for key, patrones in mapeo.items():
        for col in df.columns:
            col_upper = str(col).upper()
            for patron in patrones:
                if patron in col_upper:
                    precios[key] = col
                    break
            if key in precios:
                break
    
    if not precios and 'PRECIO' in [str(c).upper() for c in df.columns]:
        precios['P. VIP'] = 'PRECIO'
    
    return precios

def cargar_archivo(uploaded_file):
    if uploaded_file is None:
        return None
    nombre = uploaded_file.name.lower()
    try:
        if nombre.endswith('.csv'):
            try:
                df = pd.read_csv(uploaded_file, encoding='utf-8')
            except:
                df = pd.read_csv(uploaded_file, encoding='latin-1')
        else:
            df = pd.read_excel(uploaded_file)
        return limpiar_cabeceras(df)
    except Exception as e:
        st.error(f"Error: {str(e)[:80]}")
        return None

def cargar_catalogo(archivo):
    df = cargar_archivo(archivo)
    if df is None:
        return None
    return {
        'nombre': archivo.name,
        'df': df,
        'col_sku': detectar_columna_sku(df),
        'col_desc': detectar_columna_descripcion(df),
        'precios': detectar_columnas_precio(df)
    }

def buscar_stock_para_sku(sku: str, stocks: list) -> dict:
    sku_limpio = sku.strip().upper()
    stock_yessica = 0
    stock_apri004 = 0
    stock_apri001 = 0
    
    for stock in stocks:
        df = stock['df']
        df_sku = df[stock['col_sku']].astype(str).str.strip().str.upper()
        mask = df_sku == sku_limpio
        if mask.any():
            row = df[mask].iloc[0]
            col_cant = stock.get('col_cant')
            if col_cant:
                cantidad = int(corregir_numero(row[col_cant]))
                hoja = stock['hoja'].upper()
                if 'YESSICA' in hoja:
                    stock_yessica = cantidad
                elif 'APRI.004' in hoja:
                    stock_apri004 = cantidad
                elif 'APRI.001' in hoja:
                    stock_apri001 = cantidad
    
    return {
        'yessica': stock_yessica,
        'apri004': stock_apri004,
        'apri001': stock_apri001,
        'total': stock_yessica + stock_apri004 + stock_apri001
    }

def normalizar_texto(texto: str) -> str:
    if not texto or pd.isna(texto):
        return ""
    texto = texto.lower().strip()
    correcciones = {
        "xioami": "xiaomi", "xiomi": "xiaomi", "xiamoi": "xiaomi",
        "earphone": "earphone", "earphones": "earphone",
    }
    for mal, bien in correcciones.items():
        texto = texto.replace(mal, bien)
    sufijos = [' - rn', ' - es', ' - us', ' - eu', ' - gl', ' - demo', ' - rr']
    for sufijo in sufijos:
        texto = texto.replace(sufijo, '')
    return texto.strip()

def calcular_similitud(texto1: str, texto2: str) -> float:
    if not texto1 or not texto2:
        return 0.0
    texto1 = normalizar_texto(texto1)
    texto2 = normalizar_texto(texto2)
    if texto1 == texto2:
        return 100.0
    palabras1 = set(texto1.split())
    palabras2 = set(texto2.split())
    interseccion = len(palabras1.intersection(palabras2))
    union = len(palabras1.union(palabras2))
    if union == 0:
        return 0.0
    jaccard = interseccion / union
    sequence_match = SequenceMatcher(None, texto1, texto2).ratio()
    return round((jaccard * 0.6 + sequence_match * 0.4) * 100, 1)

def buscar_sku_por_descripcion(descripcion: str, catalogos: list, precio_key: str, umbral: float = 70.0):
    if not descripcion or not catalogos:
        return None
    desc_norm = normalizar_texto(descripcion)
    mejores_matches = []
    for cat in catalogos:
        df = cat['df']
        col_desc = cat.get('col_desc')
        if not col_desc:
            continue
        if precio_key not in cat.get('precios', {}):
            continue
        col_precio = cat['precios'][precio_key]
        for _, row in df.iterrows():
            desc_cat = normalizar_texto(str(row[col_desc]))
            similitud = calcular_similitud(desc_norm, desc_cat)
            if similitud >= umbral:
                try:
                    precio = float(row[col_precio]) if pd.notna(row[col_precio]) else 0
                    if precio > 0:
                        mejores_matches.append({
                            'precio': precio,
                            'sku_match': str(row[cat['col_sku']]).strip(),
                            'similitud': similitud
                        })
                except:
                    pass
    if mejores_matches:
        mejores_matches.sort(key=lambda x: x['similitud'], reverse=True)
        return mejores_matches[0]
    return None

def buscar_producto(sku: str, catalogos: list, stocks: list, precio_key: str) -> dict:
    sku_limpio = sku.strip().upper()
    stock_info = buscar_stock_para_sku(sku_limpio, stocks)
    descripcion = f"SKU: {sku}"
    precio = 0.0
    sku_equivalente = None
    similitud_equivalente = 0
    precio_equivalente = 0
    
    for cat in catalogos:
        df = cat['df']
        df_sku = df[cat['col_sku']].astype(str).str.strip().str.upper()
        mask = df_sku == sku_limpio
        if mask.any():
            row = df[mask].iloc[0]
            if precio_key in cat['precios']:
                col_precio = cat['precios'][precio_key]
                precio = corregir_numero(row[col_precio])
            if cat['col_desc']:
                descripcion = str(row[cat['col_desc']])[:200]
            break
    
    stock_total = stock_info['total']
    
    if descripcion == f"SKU: {sku}" and stock_total > 0:
        for stock in stocks:
            df = stock['df']
            df_sku = df[stock['col_sku']].astype(str).str.strip().str.upper()
            mask = df_sku == sku_limpio
            if mask.any():
                row = df[mask].iloc[0]
                for col in df.columns:
                    col_upper = str(col).upper()
                    if any(p in col_upper for p in ['DESC', 'DESCRIPCION', 'PRODUCTO', 'NOMBRE']):
                        desc_stock = str(row[col])[:200]
                        if desc_stock and desc_stock != 'nan':
                            descripcion = desc_stock
                            break
                break
    
    if precio == 0 and stock_total > 0 and descripcion and descripcion != f"SKU: {sku}":
        match = buscar_sku_por_descripcion(descripcion, catalogos, precio_key, umbral=70.0)
        if match and match['precio'] > 0:
            sku_equivalente = match['sku_match']
            similitud_equivalente = match['similitud']
            precio_equivalente = match['precio']
    
    return {
        'sku': sku,
        'descripcion': descripcion,
        'precio': precio,
        'precio_equivalente': precio_equivalente,
        'stock_yessica': stock_info['yessica'],
        'stock_apri004': stock_info['apri004'],
        'stock_apri001': stock_info['apri001'],
        'stock_total': stock_total,
        'tiene_stock': stock_total > 0,
        'tiene_precio': precio > 0,
        'sku_equivalente': sku_equivalente,
        'similitud_equivalente': similitud_equivalente
    }

def construir_badge_stock(stock_yessica, stock_apri004, stock_apri001):
    return f"""
    <div style="display:flex; flex-wrap:wrap; gap:8px; margin:8px 0;">
        <span class="badge-yessica">🟢 YESSICA: {stock_yessica}</span>
        <span class="badge-apri004">🟡 APRI.004: {stock_apri004}</span>
        <span class="badge-apri001">🔴 APRI.001: {stock_apri001}</span>
    </div>
    """

def generar_excel(items: list, cliente: str, ruc: str) -> bytes:
    output = io.BytesIO()
    df = pd.DataFrame(items)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Cotizacion', index=False, startrow=6)
        workbook = writer.book
        ws = writer.sheets['Cotizacion']
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws['A1'] = 'QTC SMART SALES PRO'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A3'] = 'FECHA:'
        ws['B3'] = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws['A4'] = 'CLIENTE:'
        ws['B4'] = cliente.upper()
        ws['A5'] = 'RUC:'
        ws['B5'] = ruc
        headers = ['SKU', 'DESCRIPCIÓN', 'CANTIDAD', 'PRECIO UNIT.', 'TOTAL']
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=7, column=i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for row_idx, item in enumerate(items, start=8):
            ws.cell(row=row_idx, column=1, value=item['sku']).border = border
            ws.cell(row=row_idx, column=2, value=item['descripcion']).border = border
            ws.cell(row=row_idx, column=3, value=item['cantidad']).border = border
            precio_cell = ws.cell(row=row_idx, column=4, value=item['precio'])
            precio_cell.number_format = '"S/." #,##0.00'
            precio_cell.border = border
            total_cell = ws.cell(row=row_idx, column=5, value=item['total'])
            total_cell.number_format = '"S/." #,##0.00'
            total_cell.border = border
        total_row = len(items) + 8
        total_label = ws.cell(row=total_row, column=4, value='TOTAL S/.')
        total_label.font = Font(bold=True, color="FFFFFF")
        total_label.fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
        total_label.border = border
        total_valor = ws.cell(row=total_row, column=5, value=sum(item['total'] for item in items))
        total_valor.number_format = '"S/." #,##0.00'
        total_valor.border = border
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 110
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.freeze_panes = 'A8'
    return output.getvalue()

# ============================================
# INICIALIZACIÓN
# ============================================

if 'auth' not in st.session_state:
    st.session_state.auth = False
if 'modo' not in st.session_state:
    st.session_state.modo = "XIAOMI"
if 'precio_key' not in st.session_state:
    st.session_state.precio_key = "P. VIP"
if 'catalogos' not in st.session_state:
    st.session_state.catalogos = []
if 'stocks' not in st.session_state:
    st.session_state.stocks = []
if 'carrito' not in st.session_state:
    st.session_state.carrito = []
if 'user_role' not in st.session_state:
    st.session_state.user_role = None
if 'user_name' not in st.session_state:
    st.session_state.user_name = None
if 'resultados_bulk' not in st.session_state:
    st.session_state.resultados_bulk = []

# ============================================
# LOGIN
# ============================================

if not st.session_state.auth:
    st.markdown("""
    <style>
        [data-testid="stSidebar"] { display: none; }
        .stApp { margin-left: 0; }
    </style>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown('<div style="background:rgba(255,255,255,0.95);border-radius:20px;padding:2rem;margin-top:50px;">', unsafe_allow_html=True)
        try:
            st.image("logo.png", width=100)
        except:
            st.markdown("<h1 style='color:#e94560;text-align:center;'>QTC</h1>", unsafe_allow_html=True)
        st.markdown("<h2 style='color:#1a1a2e;text-align:center;'>QTC Smart Sales Pro</h2>", unsafe_allow_html=True)
        st.markdown("<p style='color:#666;text-align:center;'>Sistema Profesional de Cotización</p>", unsafe_allow_html=True)
        
        usuario = st.text_input("👤 Usuario", placeholder="admin / kimberly / vendedor")
        password = st.text_input("🔒 Contraseña", type="password")
        
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("🚀 Ingresar", use_container_width=True):
                credenciales = {
                    "admin": {"password": "qtc2026", "rol": "ADMIN", "nombre": "Administrador"},
                    "kimberly": {"password": "kam2026", "rol": "KAM", "nombre": "Kimberly"},
                    "vendedor": {"password": "ventas2026", "rol": "VENDEDOR", "nombre": "Vendedor"}
                }
                if usuario in credenciales and password == credenciales[usuario]["password"]:
                    st.session_state.auth = True
                    st.session_state.user_role = credenciales[usuario]["rol"]
                    st.session_state.user_name = credenciales[usuario]["nombre"]
                    st.rerun()
                else:
                    st.error("❌ Credenciales incorrectas")
        with col_btn2:
            if st.button("👤 Invitado", use_container_width=True):
                st.session_state.auth = True
                st.session_state.user_role = "INVITADO"
                st.session_state.user_name = "Invitado"
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

# ============================================
# DESPUÉS DEL LOGIN
# ============================================

st.markdown("""
<style>
    [data-testid="stSidebar"] { display: block; }
</style>
""", unsafe_allow_html=True)

# Reconfigurar página
st.set_page_config(page_title="QTC Smart Sales Pro", page_icon="💼", layout="wide", initial_sidebar_state="expanded")

# Header
col1, col2, col3 = st.columns([1, 5, 2])
with col1:
    try:
        st.image("logo.png", width=60)
    except:
        st.markdown("**QTC**", unsafe_allow_html=True)
with col2:
    st.markdown("# QTC Smart Sales Pro")
    st.caption("Sistema Profesional de Cotización | Soporte XIAOMI · UGREEN")
with col3:
    role_badge = {"ADMIN": "🔧", "KAM": "⭐", "VENDEDOR": "🛒", "INVITADO": "👤"}
    badge = role_badge.get(st.session_state.user_role, "👤")
    st.markdown(f"""
    <div style="background: rgba(255,255,255,0.1); padding: 0.5rem 1rem; border-radius: 12px; text-align: right;">
        <span>{badge} {st.session_state.user_name}</span><br>
        <span style="font-size: 0.7rem;">{st.session_state.user_role}</span>
    </div>
    """, unsafe_allow_html=True)
    if st.button("🚪 Cerrar Sesión", key="logout"):
        st.session_state.auth = False
        st.session_state.carrito = []
        st.rerun()

st.markdown("---")

# ============================================
# SIDEBAR - CARGA DE ARCHIVOS
# ============================================

with st.sidebar:
    st.markdown("### 🎯 Configuración")
    
    marca_seleccionada = st.radio("📌 Marca / Modo", ["XIAOMI", "UGREEN"], index=0)
    st.session_state.modo = marca_seleccionada
    
    st.markdown("---")
    
    precio_opcion = st.radio("💰 Nivel de precio", ["P. VIP", "P. BOX", "P. IR"], index=0)
    st.session_state.precio_key = precio_opcion
    
    st.markdown("---")
    
    st.markdown("### 📂 Archivos")
    
    if marca_seleccionada == "XIAOMI":
        archivos_cat = st.file_uploader("📚 Catálogos", type=['xlsx', 'xls', 'csv'], accept_multiple_files=True, key="cat_upload")
        archivos_stock = st.file_uploader("📦 Stock", type=['xlsx', 'xls'], accept_multiple_files=True, key="stock_upload")
        
        if archivos_cat:
            st.session_state.catalogos = []
            for archivo in archivos_cat:
                cat = cargar_catalogo(archivo)
                if cat:
                    st.session_state.catalogos.append(cat)
                    st.success(f"✅ {archivo.name[:30]}")
        
        if archivos_stock:
            st.session_state.stocks = []
            for archivo in archivos_stock:
                try:
                    xls = pd.ExcelFile(archivo)
                    for hoja in xls.sheet_names:
                        if any(h in hoja.upper() for h in ['APRI', 'YESSICA']):
                            df = pd.read_excel(archivo, sheet_name=hoja)
                            df = limpiar_cabeceras(df)
                            col_sku = detectar_columna_sku(df)
                            col_cant = None
                            for col in df.columns:
                                if 'DISPONIBLE' in str(col).upper():
                                    col_cant = col
                                    break
                            if not col_cant:
                                for col in df.columns:
                                    if 'CANTIDAD' in str(col).upper():
                                        col_cant = col
                                        break
                            if col_cant:
                                st.session_state.stocks.append({
                                    'nombre': f"{archivo.name} [{hoja}]",
                                    'df': df,
                                    'col_sku': col_sku,
                                    'col_cant': col_cant,
                                    'hoja': hoja
                                })
                                st.success(f"✅ Stock {hoja}")
                except Exception as e:
                    st.error(f"Error: {e}")
    
    st.markdown("---")
    
    if st.session_state.carrito:
        total = sum(item.get('total', 0) for item in st.session_state.carrito)
        st.metric("Total Carrito", f"S/ {total:,.2f}")
        if st.button("🧹 Limpiar carrito", use_container_width=True):
            st.session_state.carrito = []
            st.rerun()

# ============================================
# TAB 1: MODO MASIVO
# ============================================

tab1, tab2, tab3 = st.tabs(["📦 MODO MASIVO", "🔍 BÚSQUEDA", "🛒 CARRITO"])

with tab1:
    st.markdown("### 📦 Modo Masivo")
    st.caption(f"Formato: `SKU:CANTIDAD` (uno por línea)")
    
    texto_bulk = st.text_area("", height=200, placeholder="RN0200065BK8:5\nCN0200047BK8:10")
    
    col_b1, col_b2 = st.columns(2)
    with col_b1:
        if st.button("🚀 Procesar lista", type="primary", use_container_width=True):
            if not texto_bulk:
                st.warning("Ingresa productos")
            elif not st.session_state.catalogos:
                st.warning("Carga catálogos primero")
            elif not st.session_state.stocks:
                st.warning("Carga stock primero")
            else:
                pedidos = []
                for line in texto_bulk.strip().split('\n'):
                    line = line.strip()
                    if ':' in line:
                        parts = line.split(':')
                        if len(parts) == 2:
                            try:
                                sku = parts[0].strip().upper()
                                cant = int(parts[1].strip())
                                if cant > 0:
                                    pedidos.append({'sku': sku, 'cantidad': cant})
                            except:
                                pass
                
                if pedidos:
                    with st.spinner("Procesando..."):
                        resultados = []
                        for pedido in pedidos:
                            prod = buscar_producto(pedido['sku'], st.session_state.catalogos, st.session_state.stocks, st.session_state.precio_key)
                            
                            if prod['tiene_precio'] and prod['tiene_stock']:
                                cantidad_cotizar = min(pedido['cantidad'], prod['stock_total'])
                                estado = "✅ OK"
                            elif prod['tiene_stock'] and not prod['tiene_precio']:
                                cantidad_cotizar = 0
                                estado = "⚠️ Stock sin precio"
                            elif not prod['tiene_stock'] and prod['tiene_precio']:
                                cantidad_cotizar = 0
                                estado = "📋 Solo precio, sin stock"
                            else:
                                cantidad_cotizar = 0
                                estado = "❌ No disponible"
                            
                            resultados.append({
                                **prod,
                                'cantidad_solicitada': pedido['cantidad'],
                                'cantidad_cotizar': cantidad_cotizar,
                                'estado': estado
                            })
                        
                        st.session_state.resultados_bulk = resultados
                        st.success(f"✅ Procesados {len(pedidos)} productos")
                else:
                    st.warning("No se encontraron productos válidos")
    
    with col_b2:
        if st.button("📋 Agregar al carrito", use_container_width=True):
            if st.session_state.resultados_bulk:
                agregados = 0
                for prod in st.session_state.resultados_bulk:
                    if prod['cantidad_cotizar'] > 0 and prod['tiene_precio']:
                        st.session_state.carrito.append({
                            'sku': prod['sku'],
                            'descripcion': prod['descripcion'],
                            'cantidad': prod['cantidad_cotizar'],
                            'precio': prod['precio'],
                            'total': prod['precio'] * prod['cantidad_cotizar'],
                            'stock_yessica': prod['stock_yessica'],
                            'stock_apri004': prod['stock_apri004'],
                            'stock_apri001': prod['stock_apri001']
                        })
                        agregados += 1
                st.success(f"✅ Agregados {agregados} productos")
                st.rerun()
            else:
                st.warning("Primero procesa una lista")
    
    if st.session_state.resultados_bulk:
        st.markdown("---")
        st.markdown("### 📋 Resultados")
        
        total_con_precio = sum(1 for p in st.session_state.resultados_bulk if p['tiene_precio'])
        total_con_stock = sum(1 for p in st.session_state.resultados_bulk if p['tiene_stock'])
        
        st.markdown(f"""
        <div class="counter-summary">
            <div class="counter-item">📋 Total: {len(st.session_state.resultados_bulk)}</div>
            <div class="counter-item" style="background:#4CAF50;color:white;">💰 Con precio: {total_con_precio}</div>
            <div class="counter-item">📦 Con stock: {total_con_stock}</div>
        </div>
        """, unsafe_allow_html=True)
        
        for prod in st.session_state.resultados_bulk:
            badge_stock = construir_badge_stock(prod['stock_yessica'], prod['stock_apri004'], prod['stock_apri001'])
            
            st.markdown(f"""
            <div style="background:white;border-radius:16px;padding:1rem;margin-bottom:1rem;border-left:5px solid #4CAF50;box-shadow:0 2px 4px rgba(0,0,0,0.1);">
                <div style="display:flex;justify-content:space-between;">
                    <div><strong style="color:#1a1a2e;">📦 SKU: {prod['sku']}</strong></div>
                    <div><span style="background:#2196F3;color:white;padding:2px 8px;border-radius:12px;">Solicitado: {prod['cantidad_solicitada']}</span></div>
                </div>
                <div style="margin-top:8px;color:#1a1a2e;"><strong>📝 Descripción:</strong> {prod['descripcion'][:100]}</div>
                <div style="margin-top:8px;color:#1a1a2e;">💰 Precio: <strong style="color:#e67e22;">S/ {prod['precio']:,.2f}</strong></div>
                <div>{badge_stock}</div>
                <div style="margin-top:8px;"><strong>📌 Estado:</strong> {prod['estado']}</div>
            </div>
            """, unsafe_allow_html=True)
            
            if prod.get('sku_equivalente'):
                st.markdown(f"""
                <div style="background:#E8F5E9;border-radius:12px;padding:1rem;margin:0.5rem 0;border-left:4px solid #4CAF50;">
                    <strong style="color:#2E7D32;">💡 SKU EQUIVALENTE SUGERIDO</strong><br>
                    <span style="color:#1a1a2e;"><strong>SKU:</strong> <code>{prod['sku_equivalente']}</code><br>
                    <strong>Precio:</strong> S/ {prod.get('precio_equivalente', 0):,.2f}<br>
                    <strong>Coincidencia:</strong> {prod.get('similitud_equivalente', 0):.0f}%</span>
                </div>
                """, unsafe_allow_html=True)

# ============================================
# TAB 2: BÚSQUEDA
# ============================================

with tab2:
    st.markdown("### 🔍 Búsqueda Inteligente")
    
    busqueda = st.text_input("", placeholder="Ej: RN0200065BK8 o Type-C Earphones")
    
    if busqueda and len(busqueda) >= 2:
        if st.session_state.catalogos and st.session_state.stocks:
            with st.spinner("Buscando..."):
                productos = {}
                for cat in st.session_state.catalogos:
                    df = cat['df']
                    col_sku = cat['col_sku']
                    col_desc = cat.get('col_desc')
                    
                    mask_sku = df[col_sku].astype(str).str.contains(busqueda, case=False, na=False)
                    mask_desc = pd.Series([False] * len(df))
                    if col_desc:
                        mask_desc = df[col_desc].astype(str).str.contains(busqueda, case=False, na=False)
                    mask = mask_sku | mask_desc
                    
                    for _, row in df[mask].iterrows():
                        sku = str(row[col_sku]).strip().upper()
                        descripcion = str(row[col_desc])[:200] if col_desc else f"SKU: {sku}"
                        precio = 0
                        if st.session_state.precio_key in cat['precios']:
                            col_precio = cat['precios'][st.session_state.precio_key]
                            precio = corregir_numero(row[col_precio])
                        stock_info = buscar_stock_para_sku(sku, st.session_state.stocks)
                        
                        if sku not in productos:
                            productos[sku] = {
                                'sku': sku,
                                'descripcion': descripcion,
                                'precio': precio,
                                'stock_yessica': stock_info['yessica'],
                                'stock_apri004': stock_info['apri004'],
                                'stock_apri001': stock_info['apri001'],
                                'stock_total': stock_info['total'],
                                'tiene_stock': stock_info['total'] > 0
                            }
                
                if productos:
                    st.success(f"✅ {len(productos)} productos encontrados")
                    for sku, prod in productos.items():
                        badge_stock = construir_badge_stock(prod['stock_yessica'], prod['stock_apri004'], prod['stock_apri001'])
                        st.markdown(f"""
                        <div style="background:white;border-radius:16px;padding:1rem;margin-bottom:1rem;">
                            <div><strong style="color:#1a1a2e;">📦 SKU: {prod['sku']}</strong></div>
                            <div style="color:#1a1a2e;">📝 {prod['descripcion'][:80]}</div>
                            <div style="color:#e67e22;">💰 S/ {prod['precio']:,.2f}</div>
                            <div>{badge_stock}</div>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        if prod['tiene_stock'] and prod['precio'] > 0:
                            col_cant, col_btn = st.columns([1, 2])
                            with col_cant:
                                cantidad = st.number_input("Cantidad", min_value=1, max_value=prod['stock_total'], value=1, key=f"busq_{sku}")
                            with col_btn:
                                if st.button(f"➕ Agregar", key=f"add_{sku}"):
                                    st.session_state.carrito.append({
                                        'sku': prod['sku'],
                                        'descripcion': prod['descripcion'],
                                        'cantidad': cantidad,
                                        'precio': prod['precio'],
                                        'total': prod['precio'] * cantidad,
                                        'stock_yessica': prod['stock_yessica'],
                                        'stock_apri004': prod['stock_apri004'],
                                        'stock_apri001': prod['stock_apri001']
                                    })
                                    st.success(f"✅ Agregado {cantidad}x {prod['sku']}")
                                    st.rerun()
                        st.divider()
                else:
                    st.info("No se encontraron productos")
        else:
            st.warning("Carga archivos de catálogo y stock primero")

# ============================================
# TAB 3: CARRITO
# ============================================

with tab3:
    st.markdown("### 🛒 Cotización actual")
    
    if not st.session_state.carrito:
        st.info("No hay productos en el carrito")
    else:
        for idx, item in enumerate(st.session_state.carrito):
            col1, col2, col3, col4, col5, col6 = st.columns([2, 3, 1, 1, 1, 0.5])
            with col1:
                st.write(f"**{item['sku']}**")
            with col2:
                st.write(item['descripcion'][:50])
            with col3:
                nueva_cant = st.number_input("Cant", min_value=0, value=item['cantidad'], step=1, key=f"edit_{idx}", label_visibility="collapsed")
                if nueva_cant != item['cantidad']:
                    if nueva_cant == 0:
                        st.session_state.carrito.pop(idx)
                        st.rerun()
                    else:
                        item['cantidad'] = nueva_cant
                        item['total'] = item['precio'] * nueva_cant
            with col4:
                st.write(f"S/ {item['precio']:,.2f}")
            with col5:
                st.write(f"S/ {item['total']:,.2f}")
            with col6:
                if st.button("🗑️", key=f"del_{idx}"):
                    st.session_state.carrito.pop(idx)
                    st.rerun()
            
            badge = construir_badge_stock(item.get('stock_yessica', 0), item.get('stock_apri004', 0), item.get('stock_apri001', 0))
            st.markdown(badge, unsafe_allow_html=True)
            st.divider()
        
        total_general = sum(item['total'] for item in st.session_state.carrito)
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#e94560 0%,#c73e54 100%);border-radius:12px;padding:1rem;margin:1rem 0;text-align:center;">
            <span style="color:white;font-size:1.5rem;font-weight:bold;">TOTAL: S/ {total_general:,.2f}</span>
        </div>
        """, unsafe_allow_html=True)
        
        st.markdown("### 📋 Datos del cliente")
        col_cli1, col_cli2 = st.columns(2)
        with col_cli1:
            cliente = st.text_input("Nombre del cliente", placeholder="Ej: Empresa SAC")
        with col_cli2:
            ruc = st.text_input("RUC/DNI", placeholder="Ej: 20123456789")
        
        col_exp1, col_exp2 = st.columns(2)
        with col_exp1:
            if st.button("📥 Exportar Excel", type="primary", use_container_width=True):
                if cliente:
                    items_export = [{'sku': i['sku'], 'descripcion': i['descripcion'], 'cantidad': i['cantidad'], 'precio': i['precio'], 'total': i['total']} for i in st.session_state.carrito]
                    excel = generar_excel(items_export, cliente, ruc)
                    st.download_button("💾 Descargar", data=excel, file_name=f"Cotizacion_{cliente}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx", use_container_width=True)
                    st.balloons()
                else:
                    st.warning("Ingresa el nombre del cliente")
        with col_exp2:
            if st.button("🧹 Limpiar carrito", use_container_width=True):
                st.session_state.carrito = []
                st.rerun()

# ============================================
# FOOTER
# ============================================

st.markdown("---")
st.markdown(f'<div class="footer">⚡ QTC Smart Sales Pro v5.0 | Modo: {st.session_state.modo} | {datetime.now().strftime("%Y-%m-%d %H:%M")}</div>', unsafe_allow_html=True)

# modules/cart_engine.py
# Lógica del carrito y exportación a Excel

import pandas as pd
import io
from datetime import datetime
from typing import List, Dict

def generar_excel(items: List[Dict], cliente: str, ruc: str) -> bytes:
    """Genera archivo Excel de cotización con formato profesional"""
    output = io.BytesIO()
    
    df = pd.DataFrame(items)
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Cotizacion', index=False, startrow=6)
        
        workbook = writer.book
        ws = writer.sheets['Cotizacion']
        
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados de la empresa
        ws['A1'] = 'QTC SMART SALES PRO'
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = 'FECHA:'
        ws['B3'] = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws['A4'] = 'CLIENTE:'
        ws['B4'] = cliente.upper()
        ws['A5'] = 'RUC:'
        ws['B5'] = ruc
        
        # Encabezados de la tabla
        headers = ['SKU', 'DESCRIPCIÓN', 'CANTIDAD', 'PRECIO UNIT.', 'TOTAL']
        for i, header in enumerate(headers, start=1):
            cell = ws.cell(row=7, column=i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        
        # Datos
        for row_idx, item in enumerate(items, start=8):
            ws.cell(row=row_idx, column=1, value=item['sku']).border = border
            ws.cell(row=row_idx, column=2, value=item['descripcion']).border = border
            ws.cell(row=row_idx, column=3, value=item['cantidad']).border = border
            
            precio_cell = ws.cell(row=row_idx, column=4, value=item['precio'])
            precio_cell.number_format = '"S/." #,##0.00'
            precio_cell.border = border
            
            total_cell = ws.cell(row=row_idx, column=5, value=item['total'])
            total_cell.number_format = '"S/." #,##0.00'
            total_cell.border = border
        
        # Total general
        total_row = len(items) + 8
        total_label = ws.cell(row=total_row, column=4, value='TOTAL S/.')
        total_label.font = Font(bold=True, color="FFFFFF")
        total_label.fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
        total_label.border = border
        
        total_valor = ws.cell(row=total_row, column=5, value=sum(item['total'] for item in items))
        total_valor.number_format = '"S/." #,##0.00'
        total_valor.border = border
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 110
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        
        # Congelar paneles
        ws.freeze_panes = 'A8'
    
    return output.getvalue()

def calcular_total_carrito(carrito: List[Dict]) -> float:
    """Calcula el total del carrito"""
    return sum(item.get('total', 0) for item in carrito)

def limpiar_carrito():
    """Limpia el carrito (función auxiliar)"""
    import streamlit as st
    st.session_state.carrito = []
